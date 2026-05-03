
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 60)
print("生成灏园别墅装修报价")
print("=" * 60)

# 加载数据
print("\n加载项目数据...")
with open("项目/灏园别墅_20260425/数据/房间信息.json", "r", encoding="utf-8") as f:
    room_info = json.load(f)

with open("项目/灏园别墅_20260425/数据/施工库.json", "r", encoding="utf-8") as f:
    construction_lib = json.load(f)

with open("项目/灏园别墅_20260425/数据/材料库.json", "r", encoding="utf-8") as f:
    material_lib = json.load(f)

print(f"房间数: {len(room_info['房间'])}")

# 创建Excel工作簿
wb = Workbook()
ws = wb.active
ws.title = "灏园别墅装修报价"

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bold_font = Font(bold=True)
title_font = Font(bold=True, size=16)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

current_row = 1

# 添加标题
ws.cell(row=current_row, column=1, value="灏园别墅 - 装修报价单")
ws.cell(row=current_row, column=1).font = title_font
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 1

ws.cell(row=current_row, column=1, value=f"生成日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 2

# 层高信息
ws.cell(row=current_row, column=1, value="层高信息")
ws.cell(row=current_row, column=1).font = bold_font
ws.cell(row=current_row, column=2, value="一层: 3.07m")
ws.cell(row=current_row, column=4, value="二层: 3.07m")
ws.cell(row=current_row, column=6, value="三层: 3.07m")
ws.cell(row=current_row, column=8, value="地下室: 2.5m")
current_row += 2

# === 第一部分：轻工辅料报价 ===
ws.cell(row=current_row, column=1, value="一、轻工辅料报价")
ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
ws.cell(row=current_row, column=1).fill = section_fill
current_row += 2

# 按楼层组织房间
floors = {
    "一层": [],
    "二层": [], 
    "三层": [],
    "地下室": []
}

for room in room_info['房间']:
    floor_name = room['名称'].split('-')[0]
    if floor_name in floors:
        floors[floor_name].append(room)

# 常用施工项目定义
standard_items = [
    {"name": "涂刷顶面界面剂", "unit": "平方米", "price": 3, "apply_to": "顶面"},
    {"name": "顶面批刮腻子", "unit": "平方米", "price": 20, "apply_to": "顶面"},
    {"name": "顶面带光打磨", "unit": "平方米", "price": 5, "apply_to": "顶面"},
    {"name": "顶面乳胶漆涂刷底漆", "unit": "平方米", "price": 15, "apply_to": "顶面"},
    {"name": "顶面乳胶漆涂刷面漆", "unit": "平方米", "price": 15, "apply_to": "顶面"},
    {"name": "涂刷墙面界面剂", "unit": "平方米", "price": 3, "apply_to": "墙面"},
    {"name": "墙面批刮腻子", "unit": "平方米", "price": 20, "apply_to": "墙面"},
    {"name": "墙面带光打磨", "unit": "平方米", "price": 5, "apply_to": "墙面"},
    {"name": "墙面乳胶漆涂刷底漆", "unit": "平方米", "price": 15, "apply_to": "墙面"},
    {"name": "墙面乳胶漆涂刷面漆", "unit": "平方米", "price": 15, "apply_to": "墙面"},
]

# 卫生间、厨房专用项目
wet_room_items = [
    {"name": "地面防水处理", "unit": "平方米", "price": 80, "apply_to": "地面"},
    {"name": "墙面防水处理", "unit": "平方米", "price": 70, "apply_to": "墙面"},
    {"name": "地砖铺贴", "unit": "平方米", "price": 150, "apply_to": "地面"},
    {"name": "墙砖铺贴", "unit": "平方米", "price": 140, "apply_to": "墙面"},
    {"name": "吊顶安装", "unit": "平方米", "price": 200, "apply_to": "顶面"},
]

# 普通房间地面项目
floor_items = [
    {"name": "地面找平", "unit": "平方米", "price": 60, "apply_to": "地面"},
]

# 生成施工报价
total_construction = 0
item_number = 1

for floor_name, rooms in floors.items():
    if not rooms:
        continue
        
    # 楼层标题
    ws.cell(row=current_row, column=1, value=f"{floor_name}")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    current_row += 1
    
    for room in rooms:
        room_name = room['名称'].split('-')[1]
        
        # 房间标题
        ws.cell(row=current_row, column=1, value=f"结构位置：{room_name}")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 1
        
        # 表头
        headers = ["序号", "项目名称", "单价（元）", "单位", "数量", "金额（元）", "材料说明", "工艺做法"]
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        
        # 添加施工项目
        is_wet_room = "卫生间" in room_name or "厨房" in room_name
        
        # 顶面项目
        for item in standard_items[:5]:
            qty = room['顶面面积']
            amount = item['price'] * qty
            total_construction += amount
            
            ws.cell(row=current_row, column=1, value=item_number)
            ws.cell(row=current_row, column=2, value=item['name'])
            ws.cell(row=current_row, column=3, value=item['price'])
            ws.cell(row=current_row, column=4, value=item['unit'])
            ws.cell(row=current_row, column=5, value=round(qty, 2))
            ws.cell(row=current_row, column=6, value=round(amount, 2))
            ws.cell(row=current_row, column=7, value="按标准工艺")
            ws.cell(row=current_row, column=8, value="按规范施工")
            
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [2, 7, 8]:
                    ws.cell(row=current_row, column=col).alignment = left_alignment
                else:
                    ws.cell(row=current_row, column=col).alignment = center_alignment
            
            item_number += 1
            current_row += 1
        
        # 墙面项目
        for item in standard_items[5:]:
            qty = room['墙面面积']
            amount = item['price'] * qty
            total_construction += amount
            
            ws.cell(row=current_row, column=1, value=item_number)
            ws.cell(row=current_row, column=2, value=item['name'])
            ws.cell(row=current_row, column=3, value=item['price'])
            ws.cell(row=current_row, column=4, value=item['unit'])
            ws.cell(row=current_row, column=5, value=round(qty, 2))
            ws.cell(row=current_row, column=6, value=round(amount, 2))
            ws.cell(row=current_row, column=7, value="按标准工艺")
            ws.cell(row=current_row, column=8, value="按规范施工")
            
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).border = thin_border
                if col in [2, 7, 8]:
                    ws.cell(row=current_row, column=col).alignment = left_alignment
                else:
                    ws.cell(row=current_row, column=col).alignment = center_alignment
            
            item_number += 1
            current_row += 1
        
        # 地面项目
        if is_wet_room:
            for item in wet_room_items:
                if item['apply_to'] == "地面":
                    qty = room['地面面积']
                elif item['apply_to'] == "墙面":
                    qty = room['墙面面积']
                elif item['apply_to'] == "顶面":
                    qty = room['顶面面积']
                
                amount = item['price'] * qty
                total_construction += amount
                
                ws.cell(row=current_row, column=1, value=item_number)
                ws.cell(row=current_row, column=2, value=item['name'])
                ws.cell(row=current_row, column=3, value=item['price'])
                ws.cell(row=current_row, column=4, value=item['unit'])
                ws.cell(row=current_row, column=5, value=round(qty, 2))
                ws.cell(row=current_row, column=6, value=round(amount, 2))
                ws.cell(row=current_row, column=7, value="按标准工艺")
                ws.cell(row=current_row, column=8, value="按规范施工")
                
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = thin_border
                    if col in [2, 7, 8]:
                        ws.cell(row=current_row, column=col).alignment = left_alignment
                    else:
                        ws.cell(row=current_row, column=col).alignment = center_alignment
                
                item_number += 1
                current_row += 1
        else:
            for item in floor_items:
                qty = room['地面面积']
                amount = item['price'] * qty
                total_construction += amount
                
                ws.cell(row=current_row, column=1, value=item_number)
                ws.cell(row=current_row, column=2, value=item['name'])
                ws.cell(row=current_row, column=3, value=item['price'])
                ws.cell(row=current_row, column=4, value=item['unit'])
                ws.cell(row=current_row, column=5, value=round(qty, 2))
                ws.cell(row=current_row, column=6, value=round(amount, 2))
                ws.cell(row=current_row, column=7, value="按标准工艺")
                ws.cell(row=current_row, column=8, value="按规范施工")
                
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).border = thin_border
                    if col in [2, 7, 8]:
                        ws.cell(row=current_row, column=col).alignment = left_alignment
                    else:
                        ws.cell(row=current_row, column=col).alignment = center_alignment
                
                item_number += 1
                current_row += 1
        
        current_row += 1

# === 第二部分：主材报价 ===
ws.cell(row=current_row, column=1, value="二、主材报价")
ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
ws.cell(row=current_row, column=1).fill = section_fill
current_row += 2

total_material = 0
material_item_number = 1

# 示例主材项目
material_items = [
    {"name": "客厅地砖", "unit": "平方米", "price": 350, "qty": 45.5},
    {"name": "卧室木地板", "unit": "平方米", "price": 280, "qty": 80},
    {"name": "厨房墙砖", "unit": "平方米", "price": 200, "qty": 25},
    {"name": "卫生间瓷砖", "unit": "平方米", "price": 220, "qty": 40},
    {"name": "整体橱柜", "unit": "米", "price": 2500, "qty": 6},
    {"name": "卫浴套装", "unit": "套", "price": 8000, "qty": 3},
    {"name": "室内门", "unit": "樘", "price": 2500, "qty": 10},
    {"name": "衣柜定制", "unit": "平方米", "price": 1200, "qty": 50},
    {"name": "开关插座", "unit": "项", "price": 5000, "qty": 1},
    {"name": "灯具", "unit": "项", "price": 15000, "qty": 1},
]

# 表头
material_headers = ["序号", "项目名称", "数量", "单位", "单价", "合计", "品牌", "备注"]
for col_idx, header in enumerate(material_headers):
    cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border
current_row += 1

for item in material_items:
    amount = item['price'] * item['qty']
    total_material += amount
    
    ws.cell(row=current_row, column=1, value=material_item_number)
    ws.cell(row=current_row, column=2, value=item['name'])
    ws.cell(row=current_row, column=3, value=item['qty'])
    ws.cell(row=current_row, column=4, value=item['unit'])
    ws.cell(row=current_row, column=5, value=item['price'])
    ws.cell(row=current_row, column=6, value=round(amount, 2))
    ws.cell(row=current_row, column=7, value="品牌可选")
    ws.cell(row=current_row, column=8, value="备注")
    
    for col in range(1, 9):
        ws.cell(row=current_row, column=col).border = thin_border
        if col in [2, 7, 8]:
            ws.cell(row=current_row, column=col).alignment = left_alignment
        else:
            ws.cell(row=current_row, column=col).alignment = center_alignment
    
    material_item_number += 1
    current_row += 1

current_row += 2

# === 报价汇总 ===
ws.cell(row=current_row, column=1, value="报价汇总")
ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
ws.cell(row=current_row, column=1).fill = section_fill
current_row += 2

ws.cell(row=current_row, column=2, value="轻工辅料费用:")
ws.cell(row=current_row, column=2).font = bold_font
ws.cell(row=current_row, column=6, value=round(total_construction, 2))
ws.cell(row=current_row, column=7, value="元")
current_row += 1

ws.cell(row=current_row, column=2, value="主材费用:")
ws.cell(row=current_row, column=2).font = bold_font
ws.cell(row=current_row, column=6, value=round(total_material, 2))
ws.cell(row=current_row, column=7, value="元")
current_row += 1

ws.cell(row=current_row, column=2, value="工程总报价:")
ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
ws.cell(row=current_row, column=6, value=round(total_construction + total_material, 2))
ws.cell(row=current_row, column=6).font = Font(bold=True, size=12)
ws.cell(row=current_row, column=7, value="元")
ws.cell(row=current_row, column=7).font = bold_font

# 自动调整列宽
for col in range(1, 9):
    max_length = 0
    col_letter = get_column_letter(col)
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = min(max_length + 2, 60)
    ws.column_dimensions[col_letter].width = adjusted_width

# 保存报价单
output_path = "项目/灏园别墅_20260425/报价/灏园别墅装修报价.xlsx"
wb.save(output_path)
print(f"\n报价单已保存: {output_path}")

# 保存JSON格式
quote_data = {
    "project_name": "灏园别墅",
    "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "construction_total": round(total_construction, 2),
    "material_total": round(total_material, 2),
    "grand_total": round(total_construction + total_material, 2),
    "floors": floors
}

with open("项目/灏园别墅_20260425/报价/灏园别墅装修报价.json", "w", encoding="utf-8") as f:
    json.dump(quote_data, f, ensure_ascii=False, indent=2)

print("\n" + "=" * 60)
print("报价生成完成！")
print(f"轻工辅料: {total_construction:,.2f} 元")
print(f"主材费用: {total_material:,.2f} 元")
print(f"总报价: {total_construction + total_material:,.2f} 元")
print("=" * 60)

