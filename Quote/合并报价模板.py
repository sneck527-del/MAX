
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

print("正在读取模板文件...")

# 读取主材模板
material_df = pd.read_excel("主材报价模板.xlsx", sheet_name="Sheet1")

# 读取轻工辅料模板
construction_df = pd.read_excel("轻工辅料模板.xlsx", sheet_name="Sheet1")

print(f"主材模板行数: {len(material_df)}")
print(f"轻工辅料模板行数: {len(construction_df)}")

# 解析轻工辅料模板结构
print("\n解析轻工辅料模板结构...")

# 找到标题行位置
header_row = 0
for idx, row in construction_df.iterrows():
    cell_value = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
    if "序号" in cell_value and "项目名称" in str(row.iloc[1]):
        header_row = idx
        break

print(f"表头行在第 {header_row} 行")

# 创建合并后的Excel文件
print("\n正在创建合并后的报价模板...")

wb = Workbook()
ws = wb.active
ws.title = "完整装修报价"

# 定义样式
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 当前行号
current_row = 1

# 复制轻工辅料模板的抬头信息
for idx in range(0, header_row):
    if idx < len(construction_df):
        row = construction_df.iloc[idx]
        for col_idx in range(len(row)):
            value = row.iloc[col_idx]
            if pd.notna(value):
                ws.cell(row=current_row, column=col_idx + 1, value=value)
                ws.cell(row=current_row, column=col_idx + 1).alignment = center_alignment
    current_row += 1

# 空一行
current_row += 1

# 第一部分：轻工辅料报价
print("添加轻工辅料报价部分...")

ws.cell(row=current_row, column=1, value="一、轻工辅料报价")
ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 2

# 复制轻工辅料的具体项目
section_start = header_row + 1
current_section = ""
for idx in range(section_start, len(construction_df)):
    row = construction_df.iloc[idx]
    
    # 检查是否是新的结构位置标题
    cell_0 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
    if "结构位置：" in cell_0 and pd.notna(row.iloc[5]):
        current_section = cell_0
        ws.cell(row=current_row, column=1, value=cell_0)
        ws.cell(row=current_row, column=6, value=row.iloc[5])
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        current_row += 1
        
        # 表头
        header_cells = ["序号", "项目名称", "单价（元）", "单位", "数量", "金额（元）", "材料说明", "工艺做法"]
        for col_idx, header in enumerate(header_cells):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1
        continue
    
    # 检查是否是序号行（有效数据行）
    first_cell = row.iloc[0]
    if pd.notna(first_cell) and (isinstance(first_cell, (int, float)) or str(first_cell).isdigit()):
        for col_idx in range(len(row)):
            value = row.iloc[col_idx]
            if pd.notna(value):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.alignment = left_alignment if col_idx in [1, 6, 7] else center_alignment
                cell.border = thin_border
        current_row += 1
    
    # 检查是否是空行分隔
    all_empty = True
    for col_idx in range(len(row)):
        if pd.notna(row.iloc[col_idx]):
            all_empty = False
            break
    if all_empty and current_row > 1:
        current_row += 1

# 空两行
current_row += 2

# 第二部分：主材报价
print("添加主材报价部分...")

ws.cell(row=current_row, column=1, value="二、主材报价")
ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
current_row += 2

# 解析主材模板结构
material_header_row = 0
for idx, row in material_df.iterrows():
    cell_0 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
    if "序号" in cell_0 and "项目名称" in str(row.iloc[1]):
        material_header_row = idx
        break

# 添加主材抬头信息
for idx in range(0, material_header_row):
    if idx < len(material_df):
        row = material_df.iloc[idx]
        for col_idx in range(len(row)):
            value = row.iloc[col_idx]
            if pd.notna(value):
                ws.cell(row=current_row, column=col_idx + 1, value=value)
                ws.cell(row=current_row, column=col_idx + 1).alignment = center_alignment
    current_row += 1

current_row += 1

# 添加主材项目
current_material_section = ""
for idx in range(material_header_row + 1, len(material_df)):
    row = material_df.iloc[idx]
    
    cell_0 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
    # 检查是否是阶段标题
    if "阶段" in cell_0 and pd.notna(row.iloc[0]):
        current_material_section = cell_0
        ws.cell(row=current_row, column=1, value=cell_0)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        continue
    
    # 检查是否是序号行
    first_cell = row.iloc[0]
    if pd.notna(first_cell) and (isinstance(first_cell, (int, float)) or str(first_cell).isdigit()):
        for col_idx in range(len(row)):
            value = row.iloc[col_idx]
            if pd.notna(value):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
        current_row += 1
    elif pd.notna(row.iloc[2]) and str(row.iloc[2]) != "":
        # 补充主材行（第一列为空但有数据）
        for col_idx in range(len(row)):
            value = row.iloc[col_idx]
            if pd.notna(value):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=value)
                cell.alignment = center_alignment
                cell.border = thin_border
        current_row += 1

# 自动调整列宽
for col in range(1, 13):
    max_length = 0
    col_letter = get_column_letter(col)
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = min(max_length + 2, 60)
    ws.column_dimensions[col_letter].width = adjusted_width

# 保存合并后的模板
output_filename = "完整装修报价模板.xlsx"
wb.save(output_filename)
print(f"\n✅ 合并完成！")
print(f"文件已保存: {output_filename}")

# 也保存一份JSON格式便于程序读取
print("\n正在保存JSON格式数据...")

def parse_construction_data():
    """解析施工数据"""
    data = []
    current_section = ""
    for idx in range(0, len(construction_df)):
        row = construction_df.iloc[idx]
        cell_0 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        if "结构位置：" in cell_0 and pd.notna(row.iloc[5]):
            current_section = cell_0
        elif pd.notna(row.iloc[0]) and str(row.iloc[0]).isdigit():
            item = {
                "section": current_section,
                "序号": row.iloc[0],
                "项目名称": row.iloc[1] if pd.notna(row.iloc[1]) else "",
                "单价": row.iloc[2] if pd.notna(row.iloc[2]) else 0,
                "单位": row.iloc[3] if pd.notna(row.iloc[3]) else "",
                "数量": row.iloc[4] if pd.notna(row.iloc[4]) else 0,
                "金额": row.iloc[5] if pd.notna(row.iloc[5]) else 0,
                "材料说明": row.iloc[6] if pd.notna(row.iloc[6]) else "",
                "工艺做法": row.iloc[7] if pd.notna(row.iloc[7]) else ""
            }
            data.append(item)
    return data

def parse_material_data():
    """解析主材数据"""
    data = []
    current_section = ""
    for idx in range(0, len(material_df)):
        row = material_df.iloc[idx]
        cell_0 = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        if "阶段" in cell_0 and pd.notna(row.iloc[0]):
            current_section = cell_0
        elif pd.notna(row.iloc[0]) and str(row.iloc[0]).isdigit():
            item = {
                "section": current_section,
                "序号": row.iloc[0],
                "项目名称": row.iloc[1] if pd.notna(row.iloc[1]) else "",
                "数量": row.iloc[2] if pd.notna(row.iloc[2]) else 0,
                "单位": row.iloc[3] if pd.notna(row.iloc[3]) else "",
                "单价": row.iloc[4] if pd.notna(row.iloc[4]) else 0,
                "合计": row.iloc[5] if pd.notna(row.iloc[5]) else 0,
                "说明": row.iloc[6] if pd.notna(row.iloc[6]) else "",
                "品牌": row.iloc[7] if pd.notna(row.iloc[7]) else "",
                "是否选购": row.iloc[8] if pd.notna(row.iloc[8]) else ""
            }
            data.append(item)
        elif pd.notna(row.iloc[6]) and str(row.iloc[6]) != "":
            # 补充主材
            item = {
                "section": current_section,
                "序号": None,
                "项目名称": row.iloc[6] if pd.notna(row.iloc[6]) else "",
                "数量": row.iloc[2] if pd.notna(row.iloc[2]) else 0,
                "单位": row.iloc[3] if pd.notna(row.iloc[3]) else "",
                "单价": row.iloc[4] if pd.notna(row.iloc[4]) else 0,
                "合计": row.iloc[5] if pd.notna(row.iloc[5]) else 0,
                "说明": row.iloc[6] if pd.notna(row.iloc[6]) else "",
                "品牌": row.iloc[7] if pd.notna(row.iloc[7]) else "",
                "是否选购": row.iloc[8] if pd.notna(row.iloc[8]) else ""
            }
            data.append(item)
    return data

# 保存JSON
combined_data = {
    "轻工辅料": parse_construction_data(),
    "主材": parse_material_data()
}

with open("完整报价模板.json", "w", encoding="utf-8") as f:
    json.dump(combined_data, f, ensure_ascii=False, indent=2)

print(f"JSON格式已保存: 完整报价模板.json")

